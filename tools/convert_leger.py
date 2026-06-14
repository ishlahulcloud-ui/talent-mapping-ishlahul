#!/usr/bin/env python3
"""Convert an RDM "Leger" xlsx into talent-mapping import CSVs.

Reads the non-UP class sheets (X.1, X.2, X.3, XI.1, XI.2 ...), parses the
two-level subject header, and emits CSVs matching the talent-mapping tables:
  - students.csv  : student_id(NISN), name, class, academic_year, status
  - subjects.csv  : subject_id, name, group  (core UTBK-relevant / other)
  - grades.csv    : student_id, subject_id, semester, score, source
Per-class variants (students_<class>.csv, grades_<class>.csv) are also written
to support staged, consent-gated imports.

Output goes to tools/out/ (gitignored — it contains real student PII).
Read-only on the source workbook.

Usage:
  python3 convert_leger.py "/path/to/Leger ... .xlsx" [output_dir]
"""
import csv
import os
import sys
from statistics import quantiles

import openpyxl

ACADEMIC_YEAR = "2025/2026"
SEMESTER = "Genap 2025/2026"
SOURCE = "RDM PAT 2025/2026"

# Consent basis: the enrollment form for MA Ishlahul Amanah's flagship
# "Academic Mastery & Talent Mapping" program, signed at intake, covering
# processing of grades, aptitude tests, and interviews — i.e. exactly this
# system's processing. Adjust --consent-date to the actual signing date.
# Pass --no-consent to skip generating consent.csv.
CONSENT_DATE = "2025-07-15"
CONSENT_SCOPE = "Form pendaftaran Program Academic Mastery & Talent Mapping (nilai, aptitude test, wawancara)"

# Sheets to skip: the *UP variants (no NISN), and the recap/index sheets.
def is_class_sheet(name: str) -> bool:
    n = name.strip()
    if n.upper().endswith("UP"):
        return False
    if n.lower().startswith("sheet"):
        return False
    return n[0] in ("X", "x")  # X.* and XI.*

SKIP_HEADERS = {"no", "nis", "nisn", "nama", "jk", "jumlah", "rata-rata", "rank", ""}

# UTBK-relevant subjects -> group "core".
CORE = {"BINDO", "MTK", "BING", "BIO", "KIM", "SOS", "EKO"}

SUBJECT_NAMES = {
    "QH": "Qur'an Hadis", "AA": "Akidah Akhlak", "FIK": "Fikih", "SKI": "Sejarah Kebudayaan Islam",
    "BAR": "Bahasa Arab", "PP": "Pendidikan Pancasila", "BINDO": "Bahasa Indonesia",
    "MTK": "Matematika", "IPAT": "IPA Terpadu", "IPST": "IPS Terpadu", "BING": "Bahasa Inggris",
    "PJOK": "PJOK", "INFO": "Informatika", "SB": "Seni Budaya", "SEJ": "Sejarah",
    "BIO": "Biologi", "KIM": "Kimia", "SOS": "Sosiologi", "EKO": "Ekonomi",
    "BSD": "Mulok BSD", "IKG": "Mulok IKG", "TMD": "Mulok TMD",
    "VKET": "VKET", "ELK": "ELK", "TBG": "TBG",
}


def col_map(ws):
    """Return {col_index: subject_id} from header rows 6-7, skipping id/aggregate cols."""
    r6 = [c.value for c in ws[6]]
    r7 = [c.value for c in ws[7]]
    out = {}
    for i in range(len(r6)):
        # Columns 0-4 are always the identity block (No,NIS,Nisn,Nama,JK).
        # Skip by position — a sheet may have a typo'd header here (XI.2 col1="f").
        if i < 5:
            continue
        parent = (str(r6[i]).strip() if r6[i] is not None else "")
        child = (str(r7[i]).strip() if i < len(r7) and r7[i] is not None else "")
        code = child or parent
        if not code or code.lower() in SKIP_HEADERS:
            continue
        out[i] = code
    return out


def nisn_str(v):
    if v is None:
        return ""
    if isinstance(v, float):
        v = int(v)
    if isinstance(v, int):
        return str(v).zfill(10)
    return str(v).strip()


def header_index(ws, label):
    for i, c in enumerate(ws[6]):
        if c.value is not None and str(c.value).strip().lower() == label:
            return i
    return None


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = [a for a in sys.argv[1:] if a.startswith("--")]
    if not args:
        sys.exit("usage: convert_leger.py <xlsx> [out_dir] [--consent-date=YYYY-MM-DD] [--no-consent]")
    src = args[0]
    out_dir = args[1] if len(args) > 1 else os.path.join(os.path.dirname(__file__), "out")
    os.makedirs(out_dir, exist_ok=True)

    consent_date = CONSENT_DATE
    gen_consent = "--no-consent" not in flags
    for f in flags:
        if f.startswith("--consent-date="):
            consent_date = f.split("=", 1)[1]

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    students = {}          # nisn -> (name, klass)
    dup_nisn = []          # (nisn, klass) collisions
    no_nisn = []           # (class, name)
    grades = []            # (nisn, subject_id, score)
    subjects = {}          # subject_id -> group
    per_class_students = {}
    per_class_grades = {}
    averages = []          # Rata-rata values for percentile report

    sheets = [s for s in wb.sheetnames if is_class_sheet(s)]
    for name in sheets:
        ws = wb[name]
        cmap = col_map(ws)
        for sid in cmap.values():
            subjects.setdefault(sid, "core" if sid in CORE else "other")
        nisn_i = header_index(ws, "nisn")
        nama_i = header_index(ws, "nama")
        avg_i = header_index(ws, "rata-rata")

        for row in ws.iter_rows(min_row=8, values_only=True):
            if nama_i is None or nisn_i is None:
                break
            name_v = row[nama_i] if nama_i < len(row) else None
            if not name_v or not str(name_v).strip() or str(name_v).strip().upper().startswith("LEGGER"):
                continue
            nisn = nisn_str(row[nisn_i] if nisn_i < len(row) else None)
            student_name = str(name_v).strip()
            if not nisn:
                no_nisn.append((name, student_name))
                continue
            if nisn in students:
                dup_nisn.append((nisn, name))
            else:
                students[nisn] = (student_name, name)
                per_class_students.setdefault(name, []).append((nisn, student_name))

            if avg_i is not None and avg_i < len(row) and isinstance(row[avg_i], (int, float)):
                averages.append(float(row[avg_i]))

            for ci, sid in cmap.items():
                score = row[ci] if ci < len(row) else None
                # Valid rapor score is 0-100; this also filters any id/total
                # values that leak through a malformed header.
                if isinstance(score, (int, float)) and 0 <= score <= 100:
                    sc = round(float(score), 2)
                    grades.append((nisn, sid, name, sc))
                    per_class_grades.setdefault(name, []).append((nisn, sid, sc))

    # ---- write CSVs ----
    def w(path, header, rows):
        with open(os.path.join(out_dir, path), "w", newline="", encoding="utf-8") as f:
            wr = csv.writer(f)
            wr.writerow(header)
            wr.writerows(rows)

    w("students.csv", ["student_id", "name", "class", "academic_year", "status"],
      [[nisn, nm, kl, ACADEMIC_YEAR, "active"] for nisn, (nm, kl) in students.items()])
    w("subjects.csv", ["subject_id", "name", "group"],
      [[sid, SUBJECT_NAMES.get(sid, sid), grp] for sid, grp in sorted(subjects.items())])
    w("grades.csv", ["student_id", "subject_id", "semester", "score", "source"],
      [[nisn, sid, SEMESTER, sc, SOURCE] for nisn, sid, _kl, sc in grades])

    def safe(n):
        return n.replace(".", "").replace(" ", "")
    for kl, rows in per_class_students.items():
        w(f"students_{safe(kl)}.csv", ["student_id", "name", "class", "academic_year", "status"],
          [[nisn, nm, kl, ACADEMIC_YEAR, "active"] for nisn, nm in rows])
    for kl, rows in per_class_grades.items():
        w(f"grades_{safe(kl)}.csv", ["student_id", "subject_id", "semester", "score", "source"],
          [[nisn, sid, SEMESTER, sc, SOURCE] for nisn, sid, sc in rows])

    consent_hdr = ["student_id", "student_consent_date", "parent_consent_date", "consent_scope", "withdrawal_date"]
    if gen_consent:
        w("consent.csv", consent_hdr,
          [[nisn, consent_date, consent_date, CONSENT_SCOPE, ""] for nisn in students])
        per_class_nisn = {}
        for nisn, (nm, kl) in students.items():
            per_class_nisn.setdefault(kl, []).append(nisn)
        for kl, ids in per_class_nisn.items():
            w(f"consent_{safe(kl)}.csv", consent_hdr,
              [[nisn, consent_date, consent_date, CONSENT_SCOPE, ""] for nisn in ids])

    # ---- report ----
    print(f"Sumber : {os.path.basename(src)}")
    print(f"Output : {out_dir}")
    print(f"Sheet  : {', '.join(sheets)}")
    print(f"Siswa  : {len(students)} total")
    for kl, rows in per_class_students.items():
        print(f"   {kl:6} {len(rows)} siswa")
    print(f"Mapel  : {len(subjects)} ({sum(1 for g in subjects.values() if g=='core')} core)")
    print(f"Grades : {len(grades)} baris")
    print(f"Consent: {'di-generate ('+str(len(students))+' siswa, tanggal '+consent_date+')' if gen_consent else 'dilewati (--no-consent)'}")
    print(f"NISN duplikat lintas kelas : {len(dup_nisn)} {dup_nisn[:5] if dup_nisn else ''}")
    print(f"Baris tanpa NISN           : {len(no_nisn)} {[n for _,n in no_nisn[:5]]}")
    if averages:
        qs = quantiles(averages, n=10)  # deciles
        labels = [f"p{p}" for p in range(10, 100, 10)]
        print("Distribusi Rata-rata (desil): " + ", ".join(f"{l}={v:.1f}" for l, v in zip(labels, qs)))
        print(f"   min={min(averages):.1f} max={max(averages):.1f}")


if __name__ == "__main__":
    main()
